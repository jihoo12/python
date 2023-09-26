### 파일입출략
## 파일생성
# f = open("text.txt", 'w',encoding="utf-8")
# f.write("첫 번쨰 줄 입니다")
# f.close
# f = open("text.txt", 'w',encoding="utf-8")
# for i in range(1, 11):
#     data = f"{i}번쨰 줄입니다\n"
#     f.write(data)
# f.close
## 파일 읽기
# f = open("excel.xlsx",'r',encoding='utf-8')
# for i in range(7):
#     line = f.readline()
# print(line)
# f.close

## seek
# f = open("text.txt",'r',encoding='utf-8')
# for i in range(3):
#     line = f.readline()
#     print(line)
# f.seek(67,0)
# line = f.readline(0)
# print(line)
# print(f.tell())
# # f.seek(0)
# # line = f.readline()
# # print(line)
# f.close
### 여러줄출력
## 
# f = open('text.txt','r',encoding='utf-8')
# lines = f.readlines()
# print(lines)
# f.close()
# for i in lines:
#     i = i.strip()
#     print(i,)
### 엑셀자동화
import openpyxl

# 경로 지정 (파일 이름)
save_path = 'excel.xlsx'

# 새 엑셀파일 생성
# wb = openpyxl.Workbook()
# wb.save(save_path)

## 새 시트 생성
# wb = openpyxl.Workbook()
# ws = wb.create_sheet('230926')
# wb.save(save_path)

## 시트이름 변경
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = '230925'

# wb = openpyxl.Workbook()
# ws = wb.create_sheet('230926')
# wb.save(save_path)
# wb.close()
import openpyxl

wb = openpyxl.load_workbook('excel.xlsx')
ws = wb.active
## 데이터 추가 1
ws['A1'] = '항목'
ws['B1'] = '수입'
ws['C1'] = '지출'
ws['D1'] = '잔액'
## 데이터 추가 2
sum = 0
ws.cell(row=2,column=1,value='아빠')
ws.cell(row=2,column=2,value=10000)
ws.cell(row=2,column=3,value=0)
ws.cell(row=2,column=4,value='B2')
ws.append(['간식',0,3000,'=D2-C3'])
## 데이터 수정/삭제
ws['B2'] = 1000000
ws['C3'] = 100000
del ws['A6']
wb.save('excel.xlsx')
wb.close()
